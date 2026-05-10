from django.shortcuts import render
import openpyxl
import os


def index(request):

    msg = ""

    if request.method == "POST":

        # get values from form
        master_path   = request.POST["master_path"]
        monthly_path  = request.POST["monthly_path"]
        master_sheet  = request.POST["master_sheet"]
        monthly_sheet = request.POST["monthly_sheet"]
        dedupe        = request.POST["dedupe"]

        # check if files exist
        if not os.path.exists(master_path):
            msg = "Master file not found!"

        elif not os.path.exists(monthly_path):
            msg = "Monthly file not found!"

        else:

            # open both files
            master_wb  = openpyxl.load_workbook(master_path)
            monthly_wb = openpyxl.load_workbook(monthly_path)

            # check sheet names exist
            if master_sheet not in master_wb.sheetnames:
                msg = "Sheet '" + master_sheet + "' not found in master! Available: " + ", ".join(master_wb.sheetnames)

            elif monthly_sheet not in monthly_wb.sheetnames:
                msg = "Sheet '" + monthly_sheet + "' not found in monthly! Available: " + ", ".join(monthly_wb.sheetnames)

            else:

                # get the selected sheets
                master_ws  = master_wb[master_sheet]
                monthly_ws = monthly_wb[monthly_sheet]

                # get header row from both sheets
                master_headers  = [cell.value for cell in master_ws[1]]
                monthly_headers = [cell.value for cell in monthly_ws[1]]

                # check dedupe column exists
                if dedupe not in master_headers:
                    msg = "Column " + dedupe + " not found in master file!"

                elif dedupe not in monthly_headers:
                    msg = "Column " + dedupe + " not found in monthly file!"

                else:

                    # find column index of dedupe in both sheets
                    master_col  = master_headers.index(dedupe)
                    monthly_col = monthly_headers.index(dedupe)

                    # collect all existing keys from master
                    existing_keys = set()
                    for row in master_ws.iter_rows(min_row=2, values_only=True):
                        if row[master_col] is not None:
                            existing_keys.add(str(row[master_col]).strip())

                    # add new rows from monthly to master
                    added = 0
                    for row in monthly_ws.iter_rows(min_row=2, values_only=True):

                        # get key value of this row
                        key = str(row[monthly_col]).strip() if row[monthly_col] is not None else None

                        # if key not in master then add this row
                        if key and key not in existing_keys:

                            new_row = []
                            for header in master_headers:
                                if header in monthly_headers:
                                    idx = monthly_headers.index(header)
                                    new_row.append(row[idx])
                                else:
                                    new_row.append(None)

                            master_ws.append(new_row)
                            existing_keys.add(key)
                            added += 1

                    # save directly to master file
                    master_wb.save(master_path)

                    msg = str(added) + " new rows added to master file!"

    return render(request, "index.html", {"msg": msg})


# from django.shortcuts import render
# from django.http import HttpResponse
# import openpyxl
# import xlrd
# import tempfile
# import os
# import io


# def save_temp(uploaded_file):
#     suffix = os.path.splitext(uploaded_file.name)[1]
#     tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
#     for chunk in uploaded_file.chunks():
#         tmp.write(chunk)
#     tmp.close()
#     return tmp.name


# def convert_xls_to_xlsx(path):
#     xls_wb = xlrd.open_workbook(path)
#     xls_ws = xls_wb.sheet_by_index(0)
#     new_wb = openpyxl.Workbook()
#     new_ws = new_wb.active
#     for row in range(xls_ws.nrows):
#         new_ws.append(xls_ws.row_values(row))
#     new_path = path + ".xlsx"
#     new_wb.save(new_path)
#     return new_path


# def index(request):
#     step = request.POST.get("step", "1")

#     # ── STEP 1 — receive master path + monthly file, show sheet selector ───────
#     if request.method == "POST" and step == "1":

#         master_path  = request.POST.get("master_path", "").strip()
#         monthly_file = request.FILES.get("monthly")

#         if not master_path:
#             return render(request, "index.html", {
#                 "msg": "Please enter the master file path.", "step": "1"
#             })

#         if not os.path.exists(master_path):
#             return render(request, "index.html", {
#                 "msg": f"Master file not found at: {master_path}", "step": "1"
#             })

#         if not master_path.lower().endswith((".xlsx", ".xls")):
#             return render(request, "index.html", {
#                 "msg": "Master file must be an Excel file (.xlsx or .xls)", "step": "1"
#             })

#         if not monthly_file:
#             return render(request, "index.html", {
#                 "msg": "Please upload the monthly file.", "step": "1"
#             })

#         # save monthly file to temp
#         monthly_path = save_temp(monthly_file)
#         if monthly_file.name.lower().endswith(".xls") and not monthly_file.name.lower().endswith(".xlsx"):
#             old = monthly_path
#             monthly_path = convert_xls_to_xlsx(monthly_path)
#             os.unlink(old)

#         # if master is .xls convert to temp xlsx for openpyxl to read
#         working_master   = master_path
#         master_converted = False
#         if master_path.lower().endswith(".xls") and not master_path.lower().endswith(".xlsx"):
#             working_master   = convert_xls_to_xlsx(master_path)
#             master_converted = True

#         request.session["master_path"]      = master_path
#         request.session["working_master"]   = working_master
#         request.session["master_converted"] = master_converted
#         request.session["monthly_path"]     = monthly_path

#         master_wb  = openpyxl.load_workbook(working_master)
#         monthly_wb = openpyxl.load_workbook(monthly_path)

#         return render(request, "index.html", {
#             "step"          : "2",
#             "master_path"   : master_path,
#             "master_sheets" : master_wb.sheetnames,
#             "monthly_sheets": monthly_wb.sheetnames,
#         })

#     # ── STEP 2 — merge and save directly to master path ───────────────────────
#     if request.method == "POST" and step == "2":

#         master_path      = request.session.get("master_path")
#         working_master   = request.session.get("working_master")
#         master_converted = request.session.get("master_converted", False)
#         monthly_path     = request.session.get("monthly_path")

#         master_sheet  = request.POST.get("master_sheet")
#         monthly_sheet = request.POST.get("monthly_sheet")
#         dedupe        = request.POST.get("dedupe", "").strip()

#         if not master_path or not os.path.exists(master_path):
#             return render(request, "index.html", {
#                 "msg": "Session expired or master file moved. Please start again.",
#                 "step": "1"
#             })

#         if not monthly_path or not os.path.exists(monthly_path):
#             return render(request, "index.html", {
#                 "msg": "Session expired. Please start again.",
#                 "step": "1"
#             })

#         try:
#             master_wb  = openpyxl.load_workbook(working_master)
#             monthly_wb = openpyxl.load_workbook(monthly_path)

#             master_ws  = master_wb[master_sheet]
#             monthly_ws = monthly_wb[monthly_sheet]

#             master_headers  = [cell.value for cell in master_ws[1]]
#             monthly_headers = [cell.value for cell in monthly_ws[1]]

#             if dedupe not in master_headers:
#                 return render(request, "index.html", {
#                     "msg" : f"Column '{dedupe}' not found in master. Available: {', '.join(str(h) for h in master_headers if h)}",
#                     "step": "1"
#                 })

#             if dedupe not in monthly_headers:
#                 return render(request, "index.html", {
#                     "msg" : f"Column '{dedupe}' not found in monthly. Available: {', '.join(str(h) for h in monthly_headers if h)}",
#                     "step": "1"
#                 })

#             col_master  = master_headers.index(dedupe)
#             col_monthly = monthly_headers.index(dedupe)

#             # collect existing keys
#             existing = set()
#             for row in master_ws.iter_rows(min_row=2, values_only=True):
#                 if row[col_master] is not None:
#                     existing.add(str(row[col_master]).strip())

#             added           = 0
#             missing_columns = set()

#             # append only new rows
#             for row in monthly_ws.iter_rows(min_row=2, values_only=True):
#                 key = str(row[col_monthly]).strip() if row[col_monthly] is not None else None
#                 if key and key not in existing:
#                     new_row = []
#                     for header in master_headers:
#                         if header in monthly_headers:
#                             idx = monthly_headers.index(header)
#                             new_row.append(row[idx])
#                         else:
#                             new_row.append(None)
#                             if header:
#                                 missing_columns.add(str(header))
#                     master_ws.append(new_row)
#                     existing.add(key)
#                     added += 1

#             # ── save directly back to the original master file ─────────────────
#             save_path = master_path
#             if save_path.lower().endswith(".xls") and not save_path.lower().endswith(".xlsx"):
#                 save_path = save_path + "x"  # .xls → .xlsx

#             master_wb.save(save_path)

#             msg = f"✅ {added} new rows added directly into: {save_path}"
#             if missing_columns:
#                 msg += f" | Missing columns (left blank): {', '.join(sorted(missing_columns))}"

#         except PermissionError:
#             msg = "❌ Master file is open in Excel. Please close it first and try again."

#         except Exception as e:
#             msg = f"❌ Error: {str(e)}"

#         finally:
#             _cleanup(monthly_path)
#             if master_converted and working_master != master_path:
#                 _cleanup(working_master)
#             request.session.flush()

#         return render(request, "index.html", {"msg": msg, "step": "1"})

#     # GET — show blank form
#     return render(request, "index.html", {"step": "1"})


# def _cleanup(*paths):
#     for p in paths:
#         if p and os.path.exists(p):
#             try:
#                 os.unlink(p)
#             except OSError:
#                 pass